# -*- coding: utf-8 -*-
import requests
import time
import json
from openpyxl import Workbook
import random


# 获取阅读数和点赞数
def get_more_info(link):
    # 获得mid,_biz,idx,sn 这几个在link中的信息
    mid = link.split("&")[1].split("=")[1]
    idx = link.split("&")[2].split("=")[1]
    sn = link.split("&")[3].split("=")[1]
    _biz = link.split("&")[0].split("_biz=")[1]

    # 目标url
    url = "http://mp.weixin.qq.com/mp/getappmsgext"  # 获取详情页的网址
    # 添加Cookie避免登陆操作，这里的"User-Agent"最好为手机浏览器的标识
    phone_cookie = ""
    header = {
        "Cookie": phone_cookie,
        "User-Agent": ""
    }
    # 添加data，`req_id`、`pass_ticket`分别对应文章的信息，从fiddler复制即可。
    data = {
        "is_only_read": "1",
        "is_temp_url": "0",
        "appmsg_type": "9",
        'reward_uin_count': '0'
    }
    # fillder 中取得一些不变得信息
    param_key = ""
    pass_ticket = ""  # 从fiddler中获取
    appmsg_token = ""  # 从fiddler中获取

    """
    添加请求参数
    __biz对应公众号的信息，唯一
    mid、sn、idx分别对应每篇文章的url的信息，需要从url中进行提取
    key、appmsg_token从fiddler上复制即可
    pass_ticket对应的文章的信息，也可以直接从fiddler复制
    """
    params = {
        "__biz": _biz,
        "mid": mid,
        "sn": sn,
        "idx": idx,
        "key": param_key,
        "pass_ticket": pass_ticket,
        "appmsg_token": appmsg_token,
        "uin": "", # fiddler
        "wxtoken": "777",
        "devicetype": "",
    }

    # 使用post方法进行提交
    requests.packages.urllib3.disable_warnings()
    content = requests.post(url, headers=header, data=data, params=params).json()
    # 提取其中的阅读数和点赞数
    # print(content["appmsgstat"]["read_num"], content["appmsgstat"]["like_num"])
    try:
        read_num = content["appmsgstat"]["read_num"]
        print("阅读数:" + str(read_num))
    except:
        read_num = 0
    try:
        like_num = content["appmsgstat"]["like_num"]
        print("喜爱数:" + str(like_num))
    except:
        like_num = 0
    try:
        old_like_num = content["appmsgstat"]["old_like_num"]
        print("在读数:" + str(old_like_num))
    except:
        old_like_num = 0
    # 歇3s，防止被封
    time.sleep(3)
    return read_num, like_num, old_like_num


# 微信公众号获取特定公众号文章列表的目标url
oa_url = "https://mp.weixin.qq.com/cgi-bin/appmsg"

Cookie = "you cookie"
# 使用Cookie，跳过登陆操作
headers = {
    "Cookie": Cookie,
    "User-Agent": "you agent",
}

"""
需要提交的data
以下个别字段是否一定需要还未验证。
注意修改yourtoken,number
number表示从第number页开始爬取，为5的倍数，从0开始。如0、5、10……
token可以使用Chrome自带的工具进行获取
fakeid是公众号独一无二的一个id，等同于后面的__biz
"""
token = ""  # 公众号
fakeid = ""  # 公众号对应的id
type = '9'
# 爬虫网址中的参数
data1 = {
    "token": token,
    "lang": "zh_CN",
    "f": "json",
    "ajax": "1",
    "action": "list_ex",
    "begin": "0",
    "count": "4",
    "query": "",
    "fakeid": fakeid,
    "type": type,
}


# 获取详细信息
def get_all_info(url):
    # 拿一页，存一页
    message_all_info = []
    # begin 从0开始
    for i in range(1):  # 设置爬虫页码
        begin = 500 + i * 4
        data1["begin"] = begin
        print("i:" + str(i) + " begin:" + str(begin))
        requests.packages.urllib3.disable_warnings()
        content_json = requests.get(url, headers=headers, params=data1, verify=False).json()
        if content_json['base_resp']['ret'] == 200013:
            print("frequency control, stop at {}".format(str(begin)))
            break
        time.sleep(random.randint(1, 10))
        if "app_msg_list" in content_json:
            for item in content_json["app_msg_list"]:
                spider_url = item['link']
                ten_time_array = time.localtime(item['create_time'])
                ten_other_style_time = time.strftime("%Y-%m-%d", ten_time_array)
                print("time:" + ten_other_style_time)
                read_num, like_num, old_like_num = get_more_info(spider_url)
                info = {
                    "title": item['title'],
                    "digest": item['digest'],
                    "url": item['link'],
                    "create_time": ten_other_style_time,
                    "read_num": read_num,
                    "like_num": like_num,
                    "old_like_num": old_like_num
                }
                message_all_info.append(info)
    return message_all_info


def main():
    f = Workbook()  # 创建一个workbook 设置编码
    sheet = f.active  # 创建sheet表单
    # 写入表头
    sheet.cell(row=1, column=1).value = 'title'  # 第一行第一列
    sheet.cell(row=1, column=2).value = 'digest'
    sheet.cell(row=1, column=3).value = 'url'
    sheet.cell(row=1, column=4).value = 'create_time'
    sheet.cell(row=1, column=5).value = 'readNum(阅读数)'
    sheet.cell(row=1, column=6).value = 'likeNum(喜爱数)'
    sheet.cell(row=1, column=7).value = 'old_like_num(在看数)'
    message_all_info = get_all_info(oa_url)  # 获取信息
    print(message_all_info)
    print(len(message_all_info))  # 输出列表长度
    # 写内容
    for i in range(1, len(message_all_info) + 1):
        sheet.cell(row=i + 1, column=1).value = message_all_info[i - 1]['title']
        sheet.cell(row=i + 1, column=2).value = message_all_info[i - 1]['digest']
        sheet.cell(row=i + 1, column=3).value = message_all_info[i - 1]['url']
        sheet.cell(row=i + 1, column=4).value = message_all_info[i - 1]['create_time']
        sheet.cell(row=i + 1, column=5).value = message_all_info[i - 1]['read_num']
        sheet.cell(row=i + 1, column=6).value = message_all_info[i - 1]['like_num']
        sheet.cell(row=i + 1, column=7).value = message_all_info[i - 1]['old_like_num']
    f.save(u'公众号.xlsx')  # 保存文件


if __name__ == '__main__':
    main()
